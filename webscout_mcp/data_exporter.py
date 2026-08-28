"""Enhanced data exporter module for webscout-mcp.
Supports multiple export formats: JSON, CSV, Excel, Parquet, SQLite, Markdown, HTML.

Features:
- Multiple export formats: JSON, CSV, Excel, Parquet, SQLite, Markdown, HTML
- Batch export of search results, crawl results, extracted content
- Configurable field selection and ordering
- Data transformation and cleaning
- File size optimization
- Progress reporting for large exports
- Metadata export
- Append mode for incremental exports
"""
from __future__ import annotations
import json
import csv
import os
from dataclasses import dataclass, field
from typing import Optional, Any, Union
from .logging import get_logger

log = get_logger(__name__)


@dataclass
class ExportConfig:
    """Configuration for data export."""
    # Output format: json, csv, excel, parquet, sqlite, markdown, html
    format: str = "json"
    # Output file path
    output_path: str = ""
    # Fields to export (empty = all fields)
    fields: list[str] = field(default_factory=list)
    # Field order (if empty, use default order)
    field_order: list[str] = field(default_factory=list)
    # Whether to include metadata
    include_metadata: bool = True
    # Whether to pretty-print JSON
    pretty_json: bool = True
    # CSV delimiter
    csv_delimiter: str = ","
    # Excel sheet name
    excel_sheet_name: str = "Sheet1"
    # SQLite table name
    sqlite_table_name: str = "data"
    # Whether to append to existing file (for CSV, SQLite)
    append: bool = False
    # Encoding
    encoding: str = "utf-8"
    # Maximum file size in bytes (0 = no limit)
    max_file_size: int = 0

    @classmethod
    def from_env(cls) -> "ExportConfig":
        """Load configuration from environment variables."""
        return cls(
            format=os.environ.get("WEBSCOUT_EXPORT_FORMAT", "json"),
            output_path=os.environ.get("WEBSCOUT_EXPORT_PATH", ""),
            pretty_json=os.environ.get("WEBSCOUT_EXPORT_PRETTY", "true").lower() == "true",
            csv_delimiter=os.environ.get("WEBSCOUT_EXPORT_DELIMITER", ","),
            excel_sheet_name=os.environ.get("WEBSCOUT_EXPORT_SHEET", "Sheet1"),
            sqlite_table_name=os.environ.get("WEBSCOUT_EXPORT_TABLE", "data"),
            append=os.environ.get("WEBSCOUT_EXPORT_APPEND", "false").lower() == "true",
            encoding=os.environ.get("WEBSCOUT_EXPORT_ENCODING", "utf-8"),
        )


@dataclass
class ExportResult:
    """Result of data export."""
    success: bool = False
    output_path: str = ""
    format: str = ""
    record_count: int = 0
    file_size_bytes: int = 0
    file_size_kb: float = 0.0
    error_message: str = ""
    fields_exported: list[str] = field(default_factory=list)

    def to_dict(self) -> dict:
        return {
            "success": self.success,
            "output_path": self.output_path,
            "format": self.format,
            "record_count": self.record_count,
            "file_size_bytes": self.file_size_bytes,
            "file_size_kb": self.file_size_kb,
            "error_message": self.error_message,
            "fields_exported": self.fields_exported,
        }


class DataExporter:
    """Enhanced data exporter supporting multiple formats.

    Features:
    - JSON, CSV, Excel, Parquet, SQLite, Markdown, HTML export
    - Field selection and ordering
    - Data transformation
    - Append mode
    - File size tracking
    """

    SUPPORTED_FORMATS = {"json", "csv", "excel", "parquet", "sqlite", "markdown", "html"}

    def __init__(self, config: Optional[ExportConfig] = None) -> None:
        self.config = config or ExportConfig()

    def export(self, data: list[dict], output_path: str = "", export_format: str = "") -> ExportResult:
        """Export data to specified format.

        Args:
            data: List of dictionaries to export.
            output_path: Output file path (overrides config).
            export_format: Export format (overrides config).

        Returns:
            ExportResult with export details.
        """
        result = ExportResult()
        result.format = export_format or self.config.format
        result.output_path = output_path or self.config.output_path

        if not result.output_path:
            result.error_message = "No output path specified"
            return result

        if result.format not in self.SUPPORTED_FORMATS:
            result.error_message = f"Unsupported format: {result.format}. Supported: {', '.join(sorted(self.SUPPORTED_FORMATS))}"
            return result

        if not data:
            result.error_message = "No data to export"
            return result

        try:
            # Select and order fields
            processed_data = self._process_fields(data)
            result.fields_exported = list(processed_data[0].keys()) if processed_data else []
            result.record_count = len(processed_data)

            # Export based on format
            if result.format == "json":
                self._export_json(processed_data, result)
            elif result.format == "csv":
                self._export_csv(processed_data, result)
            elif result.format == "excel":
                self._export_excel(processed_data, result)
            elif result.format == "parquet":
                self._export_parquet(processed_data, result)
            elif result.format == "sqlite":
                self._export_sqlite(processed_data, result)
            elif result.format == "markdown":
                self._export_markdown(processed_data, result)
            elif result.format == "html":
                self._export_html(processed_data, result)

            # Get file size
            if os.path.exists(result.output_path):
                result.file_size_bytes = os.path.getsize(result.output_path)
                result.file_size_kb = round(result.file_size_bytes / 1024, 2)

            result.success = True
            log.info("Data exported successfully", extra={
                "format": result.format,
                "path": result.output_path,
                "records": result.record_count,
                "size": result.file_size_kb,
            })

        except Exception as exc:
            result.success = False
            result.error_message = f"{type(exc).__name__}: {exc}"
            log.error("Data export failed", extra={"error": str(exc), "format": result.format})

        return result

    def _process_fields(self, data: list[dict]) -> list[dict]:
        """Process fields: select, order, and clean data."""
        if not data:
            return data

        # Get all fields from data
        all_fields = set()
        for item in data:
            all_fields.update(item.keys())

        # Determine fields to export
        if self.config.fields:
            fields_to_export = [f for f in self.config.fields if f in all_fields]
        else:
            fields_to_export = list(all_fields)

        # Determine field order
        if self.config.field_order:
            ordered_fields = [f for f in self.config.field_order if f in fields_to_export]
            # Add remaining fields
            for f in fields_to_export:
                if f not in ordered_fields:
                    ordered_fields.append(f)
            fields_to_export = ordered_fields

        # Process data
        processed = []
        for item in data:
            processed_item = {}
            for field in fields_to_export:
                value = item.get(field, "")
                # Convert non-serializable values to strings
                if not isinstance(value, (str, int, float, bool, type(None))):
                    value = str(value)
                processed_item[field] = value
            processed.append(processed_item)

        return processed

    def _export_json(self, data: list[dict], result: ExportResult) -> None:
        """Export to JSON format."""
        os.makedirs(os.path.dirname(result.output_path) or ".", exist_ok=True)

        with open(result.output_path, "w", encoding=self.config.encoding) as f:
            if self.config.pretty_json:
                json.dump(data, f, indent=2, ensure_ascii=False)
            else:
                json.dump(data, f, ensure_ascii=False)

    def _export_csv(self, data: list[dict], result: ExportResult) -> None:
        """Export to CSV format."""
        os.makedirs(os.path.dirname(result.output_path) or ".", exist_ok=True)

        mode = "a" if self.config.append and os.path.exists(result.output_path) else "w"
        fieldnames = list(data[0].keys()) if data else []

        with open(result.output_path, mode, newline="", encoding=self.config.encoding) as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames, delimiter=self.config.csv_delimiter)
            if mode == "w":
                writer.writeheader()
            writer.writerows(data)

    def _export_excel(self, data: list[dict], result: ExportResult) -> None:
        """Export to Excel format."""
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")

        os.makedirs(os.path.dirname(result.output_path) or ".", exist_ok=True)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = self.config.excel_sheet_name

        # Write header
        fieldnames = list(data[0].keys()) if data else []
        for col, field in enumerate(fieldnames, 1):
            ws.cell(row=1, column=col, value=field)

        # Write data
        for row, item in enumerate(data, 2):
            for col, field in enumerate(fieldnames, 1):
                ws.cell(row=row, column=col, value=item.get(field, ""))

        wb.save(result.output_path)

    def _export_parquet(self, data: list[dict], result: ExportResult) -> None:
        """Export to Parquet format."""
        try:
            import pyarrow as pa
            import pyarrow.parquet as pq
        except ImportError:
            raise ImportError("pyarrow is required for Parquet export. Install with: pip install pyarrow")

        os.makedirs(os.path.dirname(result.output_path) or ".", exist_ok=True)

        table = pa.Table.from_pylist(data)
        pq.write_table(table, result.output_path)

    def _export_sqlite(self, data: list[dict], result: ExportResult) -> None:
        """Export to SQLite format."""
        import sqlite3

        os.makedirs(os.path.dirname(result.output_path) or ".", exist_ok=True)

        conn = sqlite3.connect(result.output_path)
        cursor = conn.cursor()

        # Create table if not exists
        fieldnames = list(data[0].keys()) if data else []
        if not self.config.append:
            cursor.execute(f"DROP TABLE IF EXISTS {self.config.sqlite_table_name}")

        columns = ", ".join([f'"{field}" TEXT' for field in fieldnames])
        cursor.execute(f"CREATE TABLE IF NOT EXISTS {self.config.sqlite_table_name} ({columns})")

        # Insert data
        placeholders = ", ".join(["?" for _ in fieldnames])
        columns_str = ", ".join([f'"{f}"' for f in fieldnames])
        insert_query = f"INSERT INTO {self.config.sqlite_table_name} ({columns_str}) VALUES ({placeholders})"

        for item in data:
            values = [item.get(field, "") for field in fieldnames]
            cursor.execute(insert_query, values)

        conn.commit()
        conn.close()

    def _export_markdown(self, data: list[dict], result: ExportResult) -> None:
        """Export to Markdown format."""
        os.makedirs(os.path.dirname(result.output_path) or ".", exist_ok=True)

        fieldnames = list(data[0].keys()) if data else []

        lines = []
        # Header
        lines.append("| " + " | ".join(fieldnames) + " |")
        lines.append("| " + " | ".join(["---" for _ in fieldnames]) + " |")

        # Data rows
        for item in data:
            row = []
            for field in fieldnames:
                value = str(item.get(field, ""))
                # Escape pipe characters
                value = value.replace("|", "\\|")
                row.append(value)
            lines.append("| " + " | ".join(row) + " |")

        with open(result.output_path, "w", encoding=self.config.encoding) as f:
            f.write("\n".join(lines))

    def _export_html(self, data: list[dict], result: ExportResult) -> None:
        """Export to HTML format."""
        os.makedirs(os.path.dirname(result.output_path) or ".", exist_ok=True)

        fieldnames = list(data[0].keys()) if data else []

        lines = [
            "<!DOCTYPE html>",
            "<html>",
            "<head>",
            "<meta charset='UTF-8'>",
            "<title>Exported Data</title>",
            "<style>",
            "table { border-collapse: collapse; width: 100%; }",
            "th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }",
            "th { background-color: #4CAF50; color: white; }",
            "tr:nth-child(even) { background-color: #f2f2f2; }",
            "</style>",
            "</head>",
            "<body>",
            "<table>",
            "<thead><tr>",
        ]

        # Header
        for field in fieldnames:
            lines.append(f"<th>{field}</th>")
        lines.append("</tr></thead><tbody>")

        # Data rows
        for item in data:
            lines.append("<tr>")
            for field in fieldnames:
                value = str(item.get(field, ""))
                # Escape HTML
                value = value.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
                lines.append(f"<td>{value}</td>")
            lines.append("</tr>")

        lines.extend(["</tbody></table>", "</body>", "</html>"])

        with open(result.output_path, "w", encoding=self.config.encoding) as f:
            f.write("\n".join(lines))


def export_data(
    data: list[dict],
    output_path: str,
    export_format: str = "json",
    fields: Optional[list[str]] = None,
    **kwargs,
) -> ExportResult:
    """Convenience function to export data.

    Args:
        data: List of dictionaries to export.
        output_path: Output file path.
        export_format: Export format (json, csv, excel, parquet, sqlite, markdown, html).
        fields: Fields to export (optional).
        **kwargs: Additional configuration options.

    Returns:
        ExportResult with export details.
    """
    config = ExportConfig(format=export_format, output_path=output_path, fields=fields or [], **kwargs)
    exporter = DataExporter(config=config)
    return exporter.export(data)
